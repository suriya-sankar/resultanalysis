from django.http import HttpResponse
from .models import *
from . import forms
import openpyxl as opl
from openpyxl import load_workbook
# from openpyxl import workbook
# import pandas 
import tabula
import PyPDF2
from pandas import DataFrame
import pandas
from django.conf import settings



from tabulate import tabulate

# import camelot


from django.shortcuts import render,redirect,HttpResponseRedirect
dept='cse'

def index(request):
    batch=batches.objects.all()
    return render(request,'index.html',{'batches':batch})
    
def add_batches(request):
    if request.method=='POST':
        val=request.POST['starting_year']
        print(val)
        form=forms.Create_batches(request.POST,request.FILES)
        if form.is_valid():
            
            form.save()
            batch=batches.objects.get(starting_year=val)
            batch.batch_year_id=val[-2:]
           
            batch.save()
            dept=batch.department
            return redirect('index')
    else:        
      form=forms.Create_batches()
    return render(request,'add_batches.html',{'form':form})

def semester_list(request,batch_id):
    # return render(request,'sem.html')    
   batch_details=batches.objects.get(batch_year_id=batch_id)
#    return HttpResponse(batch_details.starting_year)
   return render(request,'sem.html',{'batch_details':batch_details}) 

def add_pdf(request,val3,batch_name,batch_id,sem_id):
   
    if request.method=='POST':
         #   semester=semester_subs(sem_id,batch_id)
           vals=request.FILES['excel_data']
           reval=0
           current_semester_excel=upload_pdf(vals,batch_id,sem_id,batch_name,reval)
        #    vals.batch_name='sjl'
        #    vals.batch_id=12
        #    vals.sem_id=3
           data=excel_raw_data()
           data.batch_name=batch_name
           data.batch_id=batch_id
           data.sem_id=sem_id
           data.excel_data=current_semester_excel
           data.save()
        #    return render(request,'sem_result_display.html',{'data':data})
           return sem_result_display(request,data,batch_id,sem_id)
    else:
       if excel_raw_data.objects.filter(batch_id=batch_id,sem_id=sem_id).exists():
         #   semester=semester_subs(sem_id,batch_id)
           data=excel_raw_data.objects.get(batch_id=batch_id,sem_id=sem_id)
        #    return render(request,'sem_result_display.html',{'data':data})
           return sem_result_display(request,data,batch_id,sem_id)
       else:    
           form=forms.upload_excel()
         #   semester=semester_subs(sem_id,batch_id)
        #    if sem_subjects.objects.filter(sem_id=sem_id,department=dept).exists():
        #         semester=sem_subjects.objects.get(sem_id=sem_id,department=dept)
        #         # return semester
        #    else:
        #        if dept is 'cse':
        #            semester=sem_subjects(1,sem_id,dept,'dd','ddd','fd','ef','gd','df','eg','cd','ds')
        #         # return semester

           return render(request,'add_pdf.html',{ 'val':val3,'batch_name':batch_name,'batch_id':batch_id,'sem_id':sem_id,'form':form});

def sem_result_display(request,data,batch_id,sem_id):
   excel_file=data.excel_data
   wb=opl.load_workbook(excel_file)
   # wb.template = True
   worksheet=wb['Sheet1']
   worksheet.delete_cols(1)
   worksheet.delete_rows(1)
  
   # data.save()
   
   excel_data_file=list()
  
   for row in worksheet.iter_rows():
     
       row_data=list()
      
       for cell in row:
           if str(cell.value)=='None' or str(cell.value)=='nan':
              
               row_data.append('    ')
           else:                     
             
               row_data.append(str(cell.value))
       excel_data_file.append(row_data)  
   m_row = worksheet.max_row 
   for i in range(1, m_row + 1): 
         cell_obj = worksheet.cell(row = i, column = 1) 
         # print(cell_obj.value) 
       
   return render(request,'sem_result_display.html',{'data':data,'excel_data_file':excel_data_file})


       
def individual_page(request,row_n,batch_id):
  
  
   register_no=row_n
   # print(register_no)
   vall=[]
   sub=[]
   batch_details=batches.objects.get(batch_year_id=batch_id)
   
   excel_data_list=excel_raw_data.objects.filter(batch_id=batch_id).order_by('sem_id')
   list_count=excel_data_list.count()
   # print(list_count)
   list_count_range=[*range(1,list_count+1)]
   # print(list_count_range)
   for db in excel_data_list:
      excel_db=db.excel_data
      wb=opl.load_workbook(excel_db)
      # print(wb)
      worksheet=wb['Sheet1']
      worksheet.delete_cols(1)
      worksheet.delete_rows(1)
      for rows in worksheet.iter_rows():
         # print(rows[1].value)
        
         if str(rows[0].value)==register_no:
            row_val=[]
            sub_val=[]
                       
            # print(str(rows[2].value))
         
            for cell in rows:
           
              if str(cell.value)=='None':
                  row_val.append('    ')
              else:
                  row_val.append(str(cell.value))
            vall.append(row_val)  
            for cell in worksheet[1]:    
              if str(cell.value)=='None':
                  sub_val.append('    ')
              else:
                 
                  sub_val.append(str(cell.value))
            sub.append(sub_val)    

    
   my_list=zip(vall,sub,list_count_range)

   return render(request,'individual_page.html',{'register_number':register_no,'batch_id':batch_id,'vall':vall,'sub':sub,'batch_details':batch_details,'range':list_count_range,'my_list':my_list})

def upload_pdf(vals,batch_id,sem_id,batch_name,reval):

   
      
      pdf_file=vals  
    
      # print(pdf_file)
      tablee=tabula.read_pdf(pdf_file,pages='all',stream=True,lattice=False)
      pdfReader = PyPDF2.PdfFileReader(pdf_file)
      pageNumber = pdfReader.numPages
      # print(pageNumber)
      sem_total_pages=list()
      sem_page=list()
   
      # print(tablee[0])
      pandas.set_option('display.precision',12)   
      tablee=DataFrame(tablee)
      table_copy=tablee
      # tablee.style.hide_index()
      
      tablee=tablee.values.tolist()
     
      # print(tablee)
      flag=1
     
      full_table_list=list()
      table_data=list()
      val=1
      lit=tablee[0].copy()
      lit=list(lit[0])
      
      for i in tablee:
         # print(val)
         current_table=i
         lit=list(i[0])
         if lit[1]=='Subject Code - >':
        
           full_table=table_data.copy()
           full_table_list.append(full_table)
           table_data.clear()
           table_data=i.copy()
         else:
        
           table_data.extend(i)  
         val=val+1
      full_table=table_data.copy()
      full_table_list.append(full_table)
      table_data.clear()   

      # print(len(full_table_list))     
      semester_id=int(sem_id)
      # print("badu")
      table_required_list=required_list(full_table_list,semester_id,batch_id) 
      # print(len(table_required_list))
      # arrear_count=arrear_count_function(table_required_list,batch_name,batch_id,sem_id,reval)
      l=len(table_required_list)
     
        
      # print(final_list) 
      if reval==1:
        for  i in range(0,semester_id):
          
         #   excel_val=reevalutation_data() 
           arrear_excel=arrear_reevaluation_data()
           mmm=DataFrame(table_required_list[i])
           j=i+1
           vall=str(j)
           bat=str(batch_id)
           se=str(semester_id)
           val="E:\excel"+"\\"+"r"+bat+se+vall+".xlsx"
           cdd="r"+bat+se+vall+".xlsx"
           exl=settings.MEDIA_ROOT+cdd
           mmm=mmm.to_excel(exl,index=False)
          
           if i == semester_id-1:
              current_excel=exl
           else:
              arrear_excel.batch_name=batch_name
              arrear_excel.batch_id=batch_id
              arrear_excel.sem_id=sem_id
              arrear_excel.arrear_sem_id=i+1
              arrear_excel.excel_data=exl
              arrear_excel.save()       
          

      else:   
         #  print("hi")
          print("hi",l)
         
          for  i in range(0,semester_id):
            # print(i)
            # print(table_required_list[i])
            # print("next")
            # excel_val=excel_raw_data() 
            arrear_excel=arrears_excel_data()
           
            mmm=DataFrame(table_required_list[i])
            j=i+1
            vall=str(j)
            bat=str(batch_id)
            se=str(semester_id)
            val="E:\excel"+"\\"+"c"+bat+se+vall+".xlsx"
            cdd="c"+bat+se+vall+".xlsx"
            exl=settings.MEDIA_ROOT+cdd
            mmm=mmm.to_excel(exl,index=False)
            if i == semester_id-1:
                 current_excel=exl
            else:
               arrear_excel.batch_name=batch_name
               arrear_excel.batch_id=batch_id
               arrear_excel.sem_id=sem_id
               arrear_excel.arrear_sem_id=i+1
               arrear_excel.excel_data=exl
               arrear_excel.save()       
       
     
      
      return current_excel
      

def required_list(full_table_list,semester_id,batch_id): 
  full_final_list=list() 
  l=len(full_table_list)
  print(l)

  for i in range(1,semester_id+1,1): 
      
      # print(i)
      oyo=list(full_table_list[i])
      # print(oyo)
      # print("hi")
      one_list=list()
      ona=list()
      final_list=list()
      another_list=list()
      yet_another_list=list()
      seco=list()
      seco_list=list()
      
      for k in oyo:
            for l in k:
              val=str(l)
            
              if val=='Subject Code - >':
                seco.append(1)
                for i in k:
                  # print(i)
                  
                   seco.append(i)
                   yet_another_list=seco.copy()

                seco_list.append(yet_another_list)
                seco.clear()
      # print(seco_list) 
               
               
         
         
      for k in oyo:
         for l in k:
            val=str(l)
            
            # if val=='Subject Code - >':
            #    # print(val)

            if val[4:6]==batch_id:
               # print(val)
               ona.append(1)
               for i in k:
                  # print(i)
                  
                  ona.append(i)
               another_list=ona.copy()
               one_list.append(another_list)
               ona.clear()
               
         # print(ona)      
         # one_list.append(ona)   
      # j=list()
      # m=list()           
      for k in oyo:
         # print(k)
         for l in k.itertuples():

            l=list(l)
            val=l[1]
            sub=l[1]
            # print(sub)
            val=str(val)
            val=val[:12]
            pos=list()
            # print(val)
           

            if val[4:6]==batch_id or val=='Reg. Number':
               # for i in l:
                  # i=str(i)
                  # j.extend(i)
               #    # m=j.copy()
              
               l[1]=val
               l=str(l)
               res = l.strip('][').split(', ')
               for c in res:
                  c=c.replace('\'','')
                  pos.append(c) 
               mark=pos.copy()
               final_list.append(mark)
               pos.clear()
            

      seco_list.extend(final_list)
      # print(one_list)
      seco_list.extend(one_list) 
      # print(final_list)   
      today_list=seco_list.copy()
      full_final_list.append(today_list)
     
      final_list.clear()
      seco_list.clear()
      # print(full_final_list)
    
  return full_final_list   

def reevalutation_upload(request,batch_id,sem_id,batch_name):

   if request.method=='POST':
      
           vals=request.FILES['add_excel']
           reval=1
           current_semester_excel=upload_pdf(vals,batch_id,sem_id,batch_name,reval)
       
           data=reevalutation_data()
           data.batch_name=batch_name
           data.batch_id=batch_id
           data.sem_id=sem_id
           data.excel_data=current_semester_excel
           data.save()
       
           return sem_result_display(request,data,batch_id,sem_id)
   else:
       if reevalutation_data.objects.filter(batch_id=batch_id,sem_id=sem_id).exists():
       
           data=reevalutation_data.objects.get(batch_id=batch_id,sem_id=sem_id)
        
           return sem_result_display(request,data,batch_id,sem_id)
       else:    
                           
           batch_details={'batch_name':batch_name,'batch_id':batch_id,'sem_id':sem_id}
           return render(request,'reevalutation_upload.html',batch_details)

def arrear_count_function(table_required_list,batch_name,batch_id,sem_id,reval):
  print("hi,arrear count funciton") 
  le=len(table_required_list)
  print(le)
  print("hi")
  sem_id=int(sem_id)



  row_list=list()
  arrear_count_ist=list()
  arrear_list=list()
  another_list=list()
  arrear_count_ist=list()
  one_another_list=list()
  if sem_id==1:
      for i in range(0,le):
        table=table_required_list[i]
        subject_row=table[0]
      #   print(subject)
        subject_len=len(subject_row)
        for rows in table:
           another_list.append(rows[1])
         #   print(rows[0][1])
           for j in range(0,subject_len):
              if rows[j]=='U':
                 row_list.append(subject_row[j])
           one_another_list=row_list.copy()      
           another_list.append(one_another_list)
           yetlist=another_list.copy()
           arrear_count_ist.append(yetlist)
           row_list.clear()
           another_list.clear()  

   #   print(arrear_count_ist)   
      arrear_count_excel=arrears_count_data()
      mmm=DataFrame(arrear_count_ist)
           
      vall=str("arrear")
      bat=str(batch_id)
           
      val="E:\excel"+"\\"+bat+vall+".xlsx"
      cdd=bat+vall+".xlsx"
      exl=settings.MEDIA_ROOT+cdd
      mmm=mmm.to_excel(exl,index=False)
          
      arrear_count_excel.batch_name=batch_name
      arrear_count_excel.batch_id=batch_id
      arrear_count_excel.excel_data=exl
      arrear_count_excel.save()       
           
              
 
  return arrear_count_ist            

        









                        








